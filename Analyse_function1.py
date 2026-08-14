import xml.etree.ElementTree as ET
from pprint import pprint
from datetime import datetime, time
import csv
import pyodbc
import openpyxl # pip install openpyxl
import winsound  # Pour Windows
# import os      # Pour Linux et macOS (nécessite 'play' ou 'afplay')


# --- Fonctions d'aide pour le parsing (inchangées) ---
def _try_parse_float(valeur_str):
    if valeur_str is None or valeur_str == '' or valeur_str == '0': return None
    try:
        valeur_nettoyee = valeur_str.replace(' ', '').replace(',', '.')
        return float(valeur_nettoyee)
    except (ValueError, TypeError): return None

def _try_parse_int(valeur_str):
    if valeur_str is None or valeur_str == '' or valeur_str == '0': return None
    try:
        valeur_nettoyee = valeur_str.replace(' ', '')
        return int(valeur_nettoyee)
    except (ValueError, TypeError): return None

def _parse_prix_achat_vente(valeur_str):
    if valeur_str is None or valeur_str == '': return None
    valeur_str = valeur_str.strip()
    if valeur_str == '0': return None
    if valeur_str.upper() in ['A', 'O', 'MARCHÉ', 'MARCHE']: return valeur_str.upper()
    return _try_parse_float(valeur_str)

def _get_champ(champs, index, defaut=''):
    try:
        return champs[index].strip() if champs[index] is not None else defaut
    except IndexError: return defaut

# --- Fonction pour parser une ligne PAC_DET et mettre à jour les caches ---
def parse_et_maj_caches_specialises(pac_det_chaine, cache_actuel, hist_prof, hist_var, hist_trans):
    pac_det_chaine = pac_det_chaine.strip()
    maj_effectuee = False

    try:
        if pac_det_chaine.startswith("~Re/!"):
            partie_donnees = pac_det_chaine.split("!", 1)[1]
            champs = partie_donnees.split('|')

            mnemonique = _get_champ(champs, 0)
            if not mnemonique:
                # print(f"AVERT: Ligne d'instrument sans mnémonique: {pac_det_chaine}") # Optionnel
                return False

            # --- Parsing des champs (identique à avant) ---
            code_isin = _get_champ(champs, 1)
            groupe_marche = _get_champ(champs, 2)
            cours_veille = _try_parse_float(_get_champ(champs, 3))
            cours_min_jour = _try_parse_float(_get_champ(champs, 4))
            cours_max_jour = _try_parse_float(_get_champ(champs, 5))
            dernier_cours = _try_parse_float(_get_champ(champs, 6))
            variation_pct_str = _get_champ(champs, 7)
            variation_pct = _try_parse_float(variation_pct_str) if variation_pct_str != '-' else 0.0
            quantite_echangee = _try_parse_int(_get_champ(champs, 8))
            seuil_bas = _try_parse_float(_get_champ(champs, 9))
            seuil_haut = _try_parse_float(_get_champ(champs, 10))
            prix_theorique = _try_parse_float(_get_champ(champs, 11))
            quantite_theorique = _try_parse_int(_get_champ(champs, 12))
            variation_theorique_str = _get_champ(champs, 13)
            cours_ouverture = _try_parse_float(_get_champ(champs, 14))

            # Carnet d'ordres
            nb_ordres_achat_1 = _try_parse_int(_get_champ(champs, 15))
            qte_achat_1 = _try_parse_int(_get_champ(champs, 16))
            cours_achat_1 = _parse_prix_achat_vente(_get_champ(champs, 17))
            cours_vente_1 = _parse_prix_achat_vente(_get_champ(champs, 18))
            qte_vente_1 = _try_parse_int(_get_champ(champs, 19))
            nb_ordres_vente_1 = _try_parse_int(_get_champ(champs, 20))
            # ... (Répéter pour les niveaux 2 à 5, ou créer une boucle)
            carnet_ordres_complet = []
            for i in range(5): # 5 niveaux
                offset = i * 6 # 6 champs par niveau (NbA, QA, CA, CV, QV, NbV)
                nb_a = _try_parse_int(_get_champ(champs, 15 + offset))
                q_a = _try_parse_int(_get_champ(champs, 16 + offset))
                c_a = _parse_prix_achat_vente(_get_champ(champs, 17 + offset))
                c_v = _parse_prix_achat_vente(_get_champ(champs, 18 + offset))
                q_v = _try_parse_int(_get_champ(champs, 19 + offset))
                nb_v = _try_parse_int(_get_champ(champs, 20 + offset))
                if any(v is not None for v in [nb_a, q_a, c_a, c_v, q_v, nb_v]): # Ajouter seulement si au moins une valeur existe
                    carnet_ordres_complet.append({
                        'nb_ordres_achat': nb_a, 'qte_achat': q_a, 'cours_achat': c_a,
                        'cours_vente': c_v, 'qte_vente': q_v, 'nb_ordres_vente': nb_v
                    })


            valeur_echangee = _try_parse_float(_get_champ(champs, 45).replace(' ', ''))
            valeur_cmp = _try_parse_float(_get_champ(champs, 46))
            nom_instrument = _get_champ(champs, 47)
            qte_dernier_echange = _try_parse_int(_get_champ(champs, 48)) # QtEE
            info_dernier_echange = _get_champ(champs, 49) # Dern
            statut_suspension = _get_champ(champs, 50).replace('\n', '').strip()
            horodatage_derniere_exec_complet = _get_champ(champs, 51) # ex: "09/05/2025 14:30:25"

            heure_derniere_exec = None
            if horodatage_derniere_exec_complet and ' ' in horodatage_derniere_exec_complet:
                 try: heure_derniere_exec = horodatage_derniere_exec_complet.split(' ')[1]
                 except IndexError: heure_derniere_exec = horodatage_derniere_exec_complet

            variation_montant = None
            if dernier_cours is not None and cours_veille is not None and cours_veille != 0:
                variation_montant = dernier_cours - cours_veille

            # --- Mettre à jour le cache actuel ---
            donnees_instrument_actuel = {
                'mnemonique': mnemonique, 'code_isin': code_isin, 'groupe_marche': groupe_marche,
                'nom': nom_instrument, 'cours_veille': cours_veille, 'cours_ouverture': cours_ouverture,
                'cours_max_jour': cours_max_jour, 'cours_min_jour': cours_min_jour, 'dernier_cours': dernier_cours,
                'variation_pourcentage': variation_pct, 'variation_montant': variation_montant,
                'quantite_echangee': quantite_echangee, 'valeur_echangee': valeur_echangee,
                'seuil_bas': seuil_bas, 'seuil_haut': seuil_haut,
                'prix_theorique_ouverture': prix_theorique, 'quantite_theorique_ouverture': quantite_theorique,
                'variation_theorique_ouverture': variation_theorique_str,
                'carnet_ordres': carnet_ordres_complet, # Stocker le carnet complet
                'valeur_cmp': valeur_cmp, 'quantite_dernier_echange': qte_dernier_echange,
                'info_dernier_echange': info_dernier_echange, 'statut_suspension': statut_suspension,
                'heure_derniere_execution': heure_derniere_exec,
                'horodatage_derniere_execution': horodatage_derniere_exec_complet
            }
            cache_actuel[mnemonique] = donnees_instrument_actuel

            # --- Clé pour l'historique : horodatage de l'exécution ou du message ---
            cle_horodatage = horodatage_derniere_exec_complet
            if not cle_horodatage: # Fallback si pas d'heure d'exécution spécifique
                cle_horodatage = cache_actuel.get('_metadata_horodatage_message', datetime.now().isoformat(sep=' ', timespec='seconds'))

            # --- 1. Historique Profondeur Marché ---
            if mnemonique not in hist_prof:
                hist_prof[mnemonique] = {}
                # On ne stocke que le carnet d'ordres et le timestamp de la source
                hist_prof[mnemonique][cle_horodatage] = {
                    'carnet_ordres': carnet_ordres_complet,
                    'horodatage_source_message': cache_actuel.get('_metadata_horodatage_message')
                }
            else :
                hist_prof[mnemonique][cle_horodatage] = {
                    'carnet_ordres': carnet_ordres_complet,
                    'horodatage_source_message': cache_actuel.get('_metadata_horodatage_message')}


            # --- 2. Historique Variation Marché ---
            if mnemonique not in hist_var:
                hist_var[mnemonique] = {}
                hist_var[mnemonique][cle_horodatage] = {
                    'dernier_cours': dernier_cours,
                    'variation_pourcentage': variation_pct,
                    'variation_montant': variation_montant,
                    'quantite_echangee_jour': quantite_echangee, # Volume total du jour
                    'valeur_echangee_jour': valeur_echangee,   # Valeur totale du jour
                    'cours_max_jour': cours_max_jour,
                    'cours_min_jour': cours_min_jour,
                    'cours_ouverture': cours_ouverture,
                    'horodatage_source_message': cache_actuel.get('_metadata_horodatage_message')
                }
            else :
                hist_var[mnemonique][cle_horodatage] = {
                    'dernier_cours': dernier_cours,
                    'variation_pourcentage': variation_pct,
                    'variation_montant': variation_montant,
                    'quantite_echangee_jour': quantite_echangee, # Volume total du jour
                    'valeur_echangee_jour': valeur_echangee,   # Valeur totale du jour
                    'cours_max_jour': cours_max_jour,
                    'cours_min_jour': cours_min_jour,
                    'cours_ouverture': cours_ouverture,
                    'horodatage_source_message': cache_actuel.get('_metadata_horodatage_message')
                }

            # --- 3. Historique Transactions (interprété comme la dernière transaction du snapshot) ---
            if qte_dernier_echange is not None and dernier_cours is not None : # Si une transaction semble avoir eu lieu
                if mnemonique not in hist_trans:
                    hist_trans[mnemonique] = {}
                    hist_trans[mnemonique][cle_horodatage] = {
                        'cours_transaction': dernier_cours, # Le dernier cours est le cours de la "dernière transaction" dans ce snapshot
                        'quantite_transaction': qte_dernier_echange, # QtEE
                        'info_transaction': info_dernier_echange, # Dern
                        'horodatage_source_message': cache_actuel.get('_metadata_horodatage_message')
                    }
                    maj_effectuee = True
                else:
                    hist_trans[mnemonique][cle_horodatage] = {
                        'cours_transaction': dernier_cours, # Le dernier cours est le cours de la "dernière transaction" dans ce snapshot
                        'quantite_transaction': qte_dernier_echange, # QtEE
                        'info_transaction': info_dernier_echange, # Dern
                        'horodatage_source_message': cache_actuel.get('_metadata_horodatage_message')
                    }
                    maj_effectuee = True


        # --- Gestion des métadonnées (pour le cache actuel) ---
        elif pac_det_chaine.startswith("~finresm/"):
            valeur = pac_det_chaine.split('/', 1)[1]
            cache_actuel['_metadata_finresm'] = _try_parse_float(valeur)
            maj_effectuee = True
        elif pac_det_chaine.startswith("~CT/"):
            valeur = pac_det_chaine.split('/', 1)[1]
            cache_actuel['_metadata_horodatage_message'] = valeur
            maj_effectuee = True
        elif pac_det_chaine.startswith("~VA/"):
            valeur = pac_det_chaine.split('/', 1)[1]
            cache_actuel['_metadata_valeur_agregee_va'] = _try_parse_int(valeur)
            maj_effectuee = True
        else:
             pass
    except Exception as e:
        print(f"ERREUR: Erreur de parsing sur la ligne: '{pac_det_chaine}' - {type(e).__name__}: {e}")
        # import traceback; traceback.print_exc() # Pour un débogage plus poussé
        return False
    return maj_effectuee

# --- Fonction pour traiter un bloc XML complet ---
def traiter_bloc_xml(chaine_xml, cache_actuel, hist_prof, hist_var, hist_trans):
    try:
        if isinstance(chaine_xml, str): chaine_xml_bytes = chaine_xml.encode('utf-8')
        else: chaine_xml_bytes = chaine_xml
        racine = ET.fromstring(chaine_xml_bytes)
        type_msg_elem = racine.find('TYPE')
        if racine.tag != 'REP' or type_msg_elem is None or type_msg_elem.text != 'MKT':
            print(f"AVERT: Format XML racine ou TYPE inattendu.")
            return False
        pacq = racine.find('PACQ')
        if pacq is None:
            print("AVERT: Balise PACQ non trouvée.")
            return False
        maj_effectuees = 0
        for pac_det in pacq.findall('PAC_DET'):
            if pac_det.text:
                if parse_et_maj_caches_specialises(pac_det.text, cache_actuel, hist_prof, hist_var, hist_trans):
                    maj_effectuees += 1
        # print(f"--- Bloc XML traité. {maj_effectuees} mises à jour détectées. Horodatage global: {cache_actuel.get('_metadata_horodatage_message', 'N/A')} ---")
        return True
    except ET.ParseError as e: print(f"ERREUR: Erreur de parsing XML: {e}"); return False
    except Exception as e: print(f"ERREUR: Erreur inattendue: {e}"); return False


#............................ Fonctions pour exporter les données ...............................

# --- Fonctions d'exportation ---

def exporter_donnees_actuelles_csv(donnees_cache, nom_fichier="donnees_actuelles.csv"):
    """Exporte le cache actuel des données de marché en CSV."""
    if not donnees_cache:
        print("Cache des données actuelles vide, rien à exporter en CSV.")
        return

    # Déterminer les en-têtes
    # On prend les clés du premier instrument comme base, en espérant qu'ils soient cohérents
    # Exclure les métadonnées commençant par '_'
    cles_instruments = [k for k in donnees_cache.keys() if not k.startswith('_')]
    if not cles_instruments:
        print("Cache des données actuelles ne contient que des métadonnées, rien à exporter en CSV.")
        return

    premier_instrument = donnees_cache[cles_instruments[0]]
    entetes_base = list(premier_instrument.keys())

    # Gérer l'aplatissement du carnet d'ordres pour CSV
    entetes_carnet = []
    max_niveaux_carnet = 0
    if 'carnet_ordres' in premier_instrument and premier_instrument['carnet_ordres']:
        max_niveaux_carnet = len(premier_instrument['carnet_ordres']) # Supposons 5
        for i in range(1, max_niveaux_carnet + 1):
            entetes_carnet.extend([
                f'carnet_nb_ordres_achat_{i}', f'carnet_qte_achat_{i}', f'carnet_cours_achat_{i}',
                f'carnet_cours_vente_{i}', f'carnet_qte_vente_{i}', f'carnet_nb_ordres_vente_{i}'
            ])
    entetes_finales = [e for e in entetes_base if e != 'carnet_ordres'] + entetes_carnet

    with open(nom_fichier, 'w', newline='', encoding='utf-8') as fichier_csv:
        writer = csv.DictWriter(fichier_csv, fieldnames=entetes_finales, extrasaction='ignore')
        writer.writeheader()
        for mnemonique, data_instrument in donnees_cache.items():
            if mnemonique.startswith('_'): # Ignorer les métadonnées
                continue
            
            ligne_aplanie = data_instrument.copy()
            if 'carnet_ordres' in ligne_aplanie:
                carnet = ligne_aplanie.pop('carnet_ordres') # Enlever la liste
                for i, niveau in enumerate(carnet):
                    if i < max_niveaux_carnet : # S'assurer de ne pas dépasser
                        ligne_aplanie[f'carnet_nb_ordres_achat_{i+1}'] = niveau.get('nb_ordres_achat')
                        ligne_aplanie[f'carnet_qte_achat_{i+1}'] = niveau.get('qte_achat')
                        ligne_aplanie[f'carnet_cours_achat_{i+1}'] = niveau.get('cours_achat')
                        ligne_aplanie[f'carnet_cours_vente_{i+1}'] = niveau.get('cours_vente')
                        ligne_aplanie[f'carnet_qte_vente_{i+1}'] = niveau.get('qte_vente')
                        ligne_aplanie[f'carnet_nb_ordres_vente_{i+1}'] = niveau.get('nb_ordres_vente')
            writer.writerow(ligne_aplanie)
    print(f"Données actuelles exportées vers {nom_fichier}")


def exporter_historique_profondeur_csv(historique, nom_fichier="hist_profondeur.csv"):
    """Exporte l'historique de la profondeur du marché en CSV."""
    if not historique:
        print("Historique de profondeur vide, rien à exporter en CSV.")
        return

    entetes = ['mnemonique', 'horodatage_execution', 'horodatage_source_message',
               'niveau_carnet', 'nb_ordres_achat', 'qte_achat', 'cours_achat',
               'cours_vente', 'qte_vente', 'nb_ordres_vente']

    with open(nom_fichier, 'w', newline='', encoding='utf-8') as fichier_csv:
        writer = csv.writer(fichier_csv)
        writer.writerow(entetes)
        for mnemonique, data_mnemo in historique.items():
            for horodatage_exec, data_exec in data_mnemo.items():
                horodatage_source = data_exec.get('horodatage_source_message', '')
                for i, niveau in enumerate(data_exec.get('carnet_ordres', [])):
                    writer.writerow([
                        mnemonique, horodatage_exec, horodatage_source, i + 1,
                        niveau.get('nb_ordres_achat'), niveau.get('qte_achat'), niveau.get('cours_achat'),
                        niveau.get('cours_vente'), niveau.get('qte_vente'), niveau.get('nb_ordres_vente')
                    ])
    print(f"Historique de profondeur exporté vers {nom_fichier}")


def exporter_historique_variation_csv(historique, nom_fichier="hist_variation.csv"):
    """Exporte l'historique des variations de marché en CSV."""
    if not historique:
        print("Historique de variation vide, rien à exporter en CSV.")
        return

    # Déterminer les en-têtes à partir du premier enregistrement
    entetes_base = []
    for mnemo_data in historique.values():
        if mnemo_data:
            first_ts_data = next(iter(mnemo_data.values()))
            entetes_base = list(first_ts_data.keys())
            break
    if not entetes_base:
        print("Aucune donnée dans l'historique de variation pour déterminer les en-têtes.")
        return

    entetes = ['mnemonique', 'horodatage_execution'] + entetes_base

    with open(nom_fichier, 'w', newline='', encoding='utf-8') as fichier_csv:
        writer = csv.DictWriter(fichier_csv, fieldnames=entetes, extrasaction='ignore')
        writer.writeheader()
        for mnemonique, data_mnemo in historique.items():
            for horodatage_exec, data_exec in data_mnemo.items():
                ligne = {'mnemonique': mnemonique, 'horodatage_execution': horodatage_exec}
                ligne.update(data_exec)
                writer.writerow(ligne)
    print(f"Historique de variation exporté vers {nom_fichier}")


def exporter_historique_transactions_csv(historique, nom_fichier="hist_transactions.csv"):
    """Exporte l'historique des transactions (interprétées) en CSV."""
    if not historique:
        print("Historique de transactions vide, rien à exporter en CSV.")
        return
    
    entetes_base = []
    for mnemo_data in historique.values():
        if mnemo_data:
            first_ts_data = next(iter(mnemo_data.values()))
            entetes_base = list(first_ts_data.keys())
            break
    if not entetes_base:
        print("Aucune donnée dans l'historique de transactions pour déterminer les en-têtes.")
        return
        
    entetes = ['mnemonique', 'horodatage_execution'] + entetes_base

    with open(nom_fichier, 'w', newline='', encoding='utf-8') as fichier_csv:
        writer = csv.DictWriter(fichier_csv, fieldnames=entetes, extrasaction='ignore')
        writer.writeheader()
        for mnemonique, data_mnemo in historique.items():
            for horodatage_exec, data_exec in data_mnemo.items():
                ligne = {'mnemonique': mnemonique, 'horodatage_execution': horodatage_exec}
                ligne.update(data_exec)
                writer.writerow(ligne)
    print(f"Historique de transactions exporté vers {nom_fichier}")

import pandas as pd


def copier_csv_vers_excel_feuille(fichier_csv_source, fichier_excel_cible, nom_feuille_cible):
    """
    Copie le contenu d'un fichier CSV dans une feuille spécifique d'un fichier Excel existant.

    Args:
        fichier_csv_source (str): Le chemin et le nom du fichier CSV source.
        fichier_excel_cible (str): Le chemin et le nom du fichier Excel cible existant.
        nom_feuille_cible (str): Le nom de la feuille dans le fichier Excel où les données seront écrites.
                                 Si la feuille n'existe pas, elle sera créée.
                                 Si elle existe, son contenu sera écrasé.
    """
    try:
        # 1. Lire le fichier CSV dans un DataFrame pandas
        df_csv = pd.read_csv(fichier_csv_source)
        print(f"Fichier CSV '{fichier_csv_source}' lu avec succès.")

        # 2. Écrire le DataFrame dans la feuille spécifique du fichier Excel existant
        # Utilisation de pd.ExcelWriter en mode 'a' (append) pour ne pas écraser le fichier
        # mais la feuille spécifiée sera écrasée si elle existe.
        with pd.ExcelWriter(fichier_excel_cible, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_csv.to_excel(writer, sheet_name=nom_feuille_cible, index=False)
        
        print(f"Données du CSV copiées avec succès vers la feuille '{nom_feuille_cible}'")
        print(f"du fichier Excel '{fichier_excel_cible}'.")

    except FileNotFoundError:
        print(f"Erreur : Le fichier CSV '{fichier_csv_source}' ou le fichier Excel '{fichier_excel_cible}' n'a pas été trouvé.")
    except Exception as e:
        print(f"Une erreur est survenue : {e}")
    

# --- Fonctions d'exportation XLSX (utilisant les fonctions CSV comme base pour les données) ---

def exporter_tout_en_xlsx(donnees_actuelles, hist_profondeur, hist_variation, hist_transactions, nom_fichier="export_complet_marche.xlsx"):
    """Exporte toutes les données dans différentes feuilles d'un fichier XLSX."""
    print("pass0")
    try:
        # Tente d'ouvrir le classeur existant
        wb = openpyxl.load_workbook(nom_fichier)
        print(f"Classeur '{nom_fichier}' ouvert avec succès.")
    except FileNotFoundError:
        # Si le fichier n'existe pas, crée un nouveau classeur
        wb = openpyxl.Workbook()
        print("Feuil créée")
        
    wb.remove(wb.active) # Supprime la feuille par défaut "Sheet"

    # Feuille 1: Données Actuelles
    if donnees_actuelles:
        print("pass3")
        ws_actuel = wb.create_sheet("Donnees Actuelles1")
        cles_instruments = [k for k in donnees_actuelles.keys() if not k.startswith('_')]
        if cles_instruments:
            premier_instrument = donnees_actuelles[cles_instruments[0]]
            entetes_base = list(premier_instrument.keys())
            entetes_carnet = []
            max_niveaux_carnet = 0
            if 'carnet_ordres' in premier_instrument and premier_instrument['carnet_ordres']:
                max_niveaux_carnet = len(premier_instrument['carnet_ordres'])
                for i in range(1, max_niveaux_carnet + 1):
                    entetes_carnet.extend([
                        f'carnet_nb_ordres_achat_{i}', f'carnet_qte_achat_{i}', f'carnet_cours_achat_{i}',
                        f'carnet_cours_vente_{i}', f'carnet_qte_vente_{i}', f'carnet_nb_ordres_vente_{i}'
                    ])
            entetes_finales = [e for e in entetes_base if e != 'carnet_ordres'] + entetes_carnet
            ws_actuel.append(entetes_finales)

            for mnemonique, data_instrument in donnees_actuelles.items():
                if mnemonique.startswith('_'): continue
                ligne_aplanie_valeurs = []
                # Construire la liste des valeurs dans le bon ordre
                temp_ligne_aplanie_dict = data_instrument.copy()
                if 'carnet_ordres' in temp_ligne_aplanie_dict:
                    carnet = temp_ligne_aplanie_dict.pop('carnet_ordres')
                    for i, niveau in enumerate(carnet):
                         if i < max_niveaux_carnet:
                            temp_ligne_aplanie_dict[f'carnet_nb_ordres_achat_{i+1}'] = niveau.get('nb_ordres_achat')
                            temp_ligne_aplanie_dict[f'carnet_qte_achat_{i+1}'] = niveau.get('qte_achat')
                            temp_ligne_aplanie_dict[f'carnet_cours_achat_{i+1}'] = niveau.get('cours_achat')
                            temp_ligne_aplanie_dict[f'carnet_cours_vente_{i+1}'] = niveau.get('cours_vente')
                            temp_ligne_aplanie_dict[f'carnet_qte_vente_{i+1}'] = niveau.get('qte_vente')
                            temp_ligne_aplanie_dict[f'carnet_nb_ordres_vente_{i+1}'] = niveau.get('nb_ordres_vente')
                
                for entete in entetes_finales:
                    ligne_aplanie_valeurs.append(temp_ligne_aplanie_dict.get(entete))
                ws_actuel.append(ligne_aplanie_valeurs)

                print(ligne_aplanie_valeurs)



    # Feuille 2: Historique Profondeur
    if hist_profondeur:
        ws_prof = wb.create_sheet("Hist Profondeur1")
        entetes = ['mnemonique', 'horodatage_execution', 'horodatage_source_message',
                   'niveau_carnet', 'nb_ordres_achat', 'qte_achat', 'cours_achat',
                   'cours_vente', 'qte_vente', 'nb_ordres_vente']
        ws_prof.append(entetes)
        for mnemonique, data_mnemo in hist_profondeur.items():
            for horodatage_exec, data_exec in data_mnemo.items():
                horodatage_source = data_exec.get('horodatage_source_message', '')
                for i, niveau in enumerate(data_exec.get('carnet_ordres', [])):
                    ws_prof.append([
                        mnemonique, horodatage_exec, horodatage_source, i + 1,
                        niveau.get('nb_ordres_achat'), niveau.get('qte_achat'), niveau.get('cours_achat'),
                        niveau.get('cours_vente'), niveau.get('qte_vente'), niveau.get('nb_ordres_vente')
                    ])

    # Feuille 3: Historique Variation
    if hist_variation:
        ws_var = wb.create_sheet("Hist Variation1")
        entetes_base = []
        for mnemo_data in hist_variation.values():
            if mnemo_data:
                first_ts_data = next(iter(mnemo_data.values()))
                entetes_base = list(first_ts_data.keys())
                break
        if entetes_base:
            entetes = ['mnemonique', 'horodatage_execution'] + entetes_base
            ws_var.append(entetes)
            for mnemonique, data_mnemo in hist_variation.items():
                for horodatage_exec, data_exec in data_mnemo.items():
                    ligne = [mnemonique, horodatage_exec] + [data_exec.get(h) for h in entetes_base]
                    ws_var.append(ligne)

    # Feuille 4: Historique Transactions
    if hist_transactions:
        ws_trans = wb.create_sheet("Hist Transactions1")
        entetes_base = []
        for mnemo_data in hist_transactions.values():
            if mnemo_data:
                first_ts_data = next(iter(mnemo_data.values()))
                entetes_base = list(first_ts_data.keys())
                break
        if entetes_base:
            entetes = ['mnemonique', 'horodatage_execution'] + entetes_base
            ws_trans.append(entetes)
            for mnemonique, data_mnemo in hist_transactions.items():
                for horodatage_exec, data_exec in data_mnemo.items():
                    ligne = [mnemonique, horodatage_exec] + [data_exec.get(h) for h in entetes_base]
                    ws_trans.append(ligne)
    
    if wb.sheetnames: # S'assurer qu'il y a au moins une feuille avant de sauvegarder
        wb.save(nom_fichier)
        print(f"Toutes les données exportées vers {nom_fichier}")
    else:
        print("Aucune donnée à exporter dans le fichier XLSX.")

# ............................................. Fonctions pour l'insertion de données dans dans une base de données SQL ................
# Création de table:

def NewTable(table_name,database, cnxn_str) :

    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # Requête SQL pour créer la table
        # Il est bon de vérifier si la table existe avant de la créer
        check_table_sql = f"SELECT OBJECT_ID('{table_name}', 'U')"
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                ID INT IDENTITY(1,1) PRIMARY KEY,
                Nom VARCHAR(255) NOT NULL,
                Email VARCHAR(255) UNIQUE,
                DateCreation DATETIME DEFAULT GETDATE()
            );
            """
            cursor.execute(create_table_sql)
            cnxn.commit() # Valider la création de la table
            print(f"Table '{table_name}' créée avec succès ! ✅")
        else:
            print(f"La table '{table_name}' existe déjà. ⚠️")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la création de la table : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")

# Insertion de valeurs dans une table d'une base de donnée

def InsertData(table_name,database, cnxn_str):
    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # --- Liste de listes ou de tuples de données à insérer ---
        # Chaque sous-liste/tuple représente une ligne, et les éléments
        # correspondent aux colonnes dans l'ordre spécifié dans la requête SQL.
        i= 1
        data_to_insert = []
        while i==1 :
            nom = input("\n Nom : ")
            email = input("\n e-mail : ")
            data_to_insert.append((nom,email))
            i=int(input("Voulez vous entrez une nouvelle donnée? si oui entre 1 si non 0 : "))
            
        # --- Requête SQL d'insertion paramétrée (la même que pour une seule ligne) ---
        insert_sql = f"INSERT INTO {table_name} (Nom, Email) VALUES (?, ?)"

        # --- Exécuter la requête pour plusieurs lignes ---
        # pyodbc prend la requête et une séquence de séquences (liste de tuples/listes)
        cursor.executemany(insert_sql, data_to_insert)

        # --- Valider les changements ---
        cnxn.commit()
        print(f"{len(data_to_insert)} lignes insérées avec succès via executemany ! ✅")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de l'insertion de données : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()
        print("Opération d'insertion annulée. ↩️")

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")







def InsertActualData(table_name,database, cnxn_str,donnees_actuelles):
    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        data_to_insert = []
        if donnees_actuelles:
            #ws_actuel = wb.create_sheet("Donnees Actuelles1")
            cles_instruments = [k for k in donnees_actuelles.keys() if not k.startswith('_')]
            if cles_instruments:
                premier_instrument = donnees_actuelles[cles_instruments[0]]
                entetes_base = list(premier_instrument.keys())
                entetes_carnet = []
                max_niveaux_carnet = 0
                if 'carnet_ordres' in premier_instrument and premier_instrument['carnet_ordres']:
                    max_niveaux_carnet = len(premier_instrument['carnet_ordres'])
                    for i in range(1, max_niveaux_carnet + 1):
                        entetes_carnet.extend([
                            f'carnet_nb_ordres_achat_{i}', f'carnet_qte_achat_{i}', f'carnet_cours_achat_{i}',
                            f'carnet_cours_vente_{i}', f'carnet_qte_vente_{i}', f'carnet_nb_ordres_vente_{i}'
                        ])
                entetes_finales = [e for e in entetes_base if e != 'carnet_ordres'] + entetes_carnet
                #ws_actuel.append(entetes_finales)

                for mnemonique, data_instrument in donnees_actuelles.items():
                    if mnemonique.startswith('_'): continue
                    ligne_aplanie_valeurs = []
                    # Construire la liste des valeurs dans le bon ordre
                    temp_ligne_aplanie_dict = data_instrument.copy()
                    if 'carnet_ordres' in temp_ligne_aplanie_dict:
                        carnet = temp_ligne_aplanie_dict.pop('carnet_ordres')
                        for i, niveau in enumerate(carnet):
                            if i < max_niveaux_carnet:
                                temp_ligne_aplanie_dict[f'carnet_nb_ordres_achat_{i+1}'] = niveau.get('nb_ordres_achat')
                                temp_ligne_aplanie_dict[f'carnet_qte_achat_{i+1}'] = niveau.get('qte_achat')
                                temp_ligne_aplanie_dict[f'carnet_cours_achat_{i+1}'] = niveau.get('cours_achat')
                                temp_ligne_aplanie_dict[f'carnet_cours_vente_{i+1}'] = niveau.get('cours_vente')
                                temp_ligne_aplanie_dict[f'carnet_qte_vente_{i+1}'] = niveau.get('qte_vente')
                                temp_ligne_aplanie_dict[f'carnet_nb_ordres_vente_{i+1}'] = niveau.get('nb_ordres_vente')
                    
                    for entete in entetes_finales:
                        ligne_aplanie_valeurs.append(temp_ligne_aplanie_dict.get(entete))
                    #ws_actuel.append(ligne_aplanie_valeurs)

                    data_to_insert.append(list(ligne_aplanie_valeurs))
            
        # 1. Définir les bornes de l'intervalle de collecte de données et l'heure actuelle
        heure_debut = time(10, 45) 
        heure_fin = time(18, 5)    

        heure_actuelle = datetime.now().time()
        if heure_debut <= heure_actuelle <= heure_fin :

            # --- Requête SQL d'insertion paramétrée (la même que pour une seule ligne) ---
            insert_sql = f"INSERT INTO {table_name} (mnemonique, code_isin, groupe_marche, Nom, cours_veille, cours_ouverture, cours_max_jour, cours_min_jour, dernier_cours, variation_pourcentage, variation_montant, quantite_echangee, valeur_echangee, seuil_bas, seuil_haut, prix_theorique_ouverture, quantite_theorique_ouverture, variation_theorique_ouverture, valeur_cmp, quantite_dernier_echange, info_dernier_echange, statut_suspension, heure_derniere_execution, horodatage_derniere_execution) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"

            # --- Exécuter la requête pour plusieurs lignes ---
            # pyodbc prend la requête et une séquence de séquences (liste de tuples/listes)
            cursor.executemany(insert_sql, data_to_insert)

            # --- Valider les changements ---
            cnxn.commit()
            print(f"{len(data_to_insert)} lignes insérées avec succès via executemany ! ✅")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de l'insertion de données : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()
        print("Opération d'insertion annulée. ↩️")

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")


def vider_table_de_donnees(table_name, cnxn_str, db_name):
    # --- Nom de la table dont vous voulez effacer les données ---
    table_a_vider = table_name # <--- Remplacez par le nom de votre table !

    cnxn = None
    cursor = None

    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{db_name}'.")

        # Exécuter la commande TRUNCATE TABLE
        # Il est important d'utiliser le format [schema].[nom_table] pour éviter les ambiguïtés
        truncate_sql = f"TRUNCATE TABLE [dbo].[{table_a_vider}];" # Assurez-vous que 'dbo' est le bon schéma
        cursor.execute(truncate_sql)

        cnxn.commit() # Valider la transaction
        print(f"Toutes les données de la table '{table_a_vider}' ont été effacées avec succès via TRUNCATE TABLE ! ✅")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors du vidage de la table '{table_a_vider}' : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback() # Annuler les modifications en cas d'erreur
        print("Opération de vidage annulée. ↩️")
    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e} ❌")
    finally:
        if cursor:
            cursor.close()
        if cnxn:
            cnxn.close()
            print("Connexion fermée. 🔒")
            

def Collecte_data_BRVM(donnees_marche_cache_actuellement):

    # Détails de la connexion
    # Remplacez les valeurs entre crochets <> par les vôtres
    server = 'DESKTOP-ME6GROA\SQLSERVER12' # Ex: 'localhost\SQLEXPRESS' ou 'VOTRE_IP'
    database = 'BRVMDATA'
    #username = 'DESKTOP-HP-LUCA\Luca'
    #password = ''

    #f"DRIVER={{ODBC Driver 17 for SQL Server}};" 
    # Chaîne de connexion
    cnxn_str = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"Trusted_Connection=yes;" # Indique d'utiliser l'authentification Windows
    )


    try:
        #Tache = int(input("Sélectionner votre tâche :\n 1- Création de table\n 2- Insertion de données dans une table \n 3- Suppression des données de la table des données (Donnees_actuelle)"))
        Tache = 2
        if Tache == 1 :
            table_name = "DataActualBRVMConbine" # Le nom de la table à créer
            ActualDataTable(table_name, database, cnxn_str)
        elif Tache == 2 :
            Database = "BRVMDATA"
            table_name = "DataActualBRVMConbine" 
            InsertActualData(table_name, Database, cnxn_str, donnees_marche_cache_actuellement)
        elif Tache == 3 :
            Database = "BRVMDATA"
            table_name = "DataActualBRVMConbine" 
            vider_table_de_donnees(table_name, cnxn_str, Database)
        else :
            pass

        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print("Connexion établie avec succès avec authentification Windows ! 🎉")
        """
        i=0
        while i ==0 :
            i = int(input(" choisis un nombre différent de 0 pour continuer les , sinon choisis 0"))
            cnxn = pyodbc.connect(cnxn_str)
            cursor = cnxn.cursor()
            print("Connexion établie avec succès avec authentification Windows ! 🎉")
            # ... votre code pour insérer des données ...
        """
            

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur de connexion : {sqlstate}")
        print(ex)
    finally:
        if 'cursor' in locals() and cursor is not None:
            cursor.close()
        if 'cnxn' in locals() and cnxn is not None:
            cnxn.close()
            print("Connexion fermée. 🔒")

def ActualDataTable(table_name,database, cnxn_str) :

    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # Requête SQL pour créer la table
        # Il est bon de vérifier si la table existe avant de la créer
        check_table_sql = f"SELECT OBJECT_ID('{table_name}', 'U')"
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                ID INT IDENTITY(1,1) PRIMARY KEY,
                mnemonique VARCHAR(255) NOT NULL,
                code_isin VARCHAR(255) NOT NULL,
                groupe_marche INT NOT NULL,
                Nom VARCHAR(255) NOT NULL,
                cours_veille FLOAT(6) ,
                cours_ouverture FLOAT(6) ,
                cours_max_jour FLOAT(6) ,
                cours_min_jour FLOAT(6) ,
                dernier_cours FLOAT(6) ,
                variation_pourcentage FLOAT(6) ,
                variation_montant FLOAT(6) ,
                quantite_echangee INT ,
                valeur_echangee INT ,
                seuil_bas INT ,
                seuil_haut INT ,
                prix_theorique_ouverture FLOAT(10) ,
                quantite_theorique_ouverture FLOAT(10) ,
                variation_theorique_ouverture INT,
                valeur_cmp FLOAT(6),
                quantite_dernier_echange INT,
                info_dernier_echange FLOAT(6),
                statut_suspension INT,
                heure_derniere_execution TIME,
                horodatage_derniere_execution DATETIME,
                DateCreation DATETIME DEFAULT GETDATE()
            );
            """
            cursor.execute(create_table_sql)
            cnxn.commit() # Valider la création de la table
            print(f"Table '{table_name}' créée avec succès ! ✅")
        else:
            print(f"La table '{table_name}' existe déjà. ⚠️")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la création de la table : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")

def Client(table_name,database, cnxn_str) :

    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # Requête SQL pour créer la table
        # Il est bon de vérifier si la table existe avant de la créer
        check_table_sql = f"SELECT OBJECT_ID('{table_name}', 'U')"
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                ID_Client INT IDENTITY(1,1) PRIMARY KEY,
                Numero VARCHAR(255) NOT NULL,
                Nom VARCHAR(255) NOT NULL,
                Prenoms VARCHAR(255) NOT NULL,
                Nationalite VARCHAR(255) NOT NULL,
                Proflit VARCHAR(255) NOT NULL,
                DateCreation DATETIME DEFAULT GETDATE(),
            );
            """
            cursor.execute(create_table_sql)
            cnxn.commit() # Valider la création de la table
            print(f"Table '{table_name}' créée avec succès ! ✅")
        else:
            print(f"La table '{table_name}' existe déjà. ⚠️")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la création de la table : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")

def Transaction(table_name,database, cnxn_str) :

    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # Requête SQL pour créer la table
        # Il est bon de vérifier si la table existe avant de la créer
        check_table_sql = f"SELECT OBJECT_ID('{table_name}', 'U')"
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                ID_Transaction INT IDENTITY(1,1) PRIMARY KEY,
                ID_Client INT NOT NULL,
                DateTransaction DATETIME DEFAULT GETDATE(),
                Numero_Ordre INT NOT NULL,
                Type_Transaction VARCHAR(255) NOT NULL,
                mnemonique VARCHAR(255) NOT NULL,
                Quantite INT NOT NULL,
                Debit INT NOT NULL,
                Credit INT NOT NULL,
                Pays VARCHAR(255) NOT NULL,
                Secteur VARCHAR(255) NOT NULL,

                -- Clé secondaire
                FOREIGN KEY (ID_Client) REFERENCES Client(ID_Client)
            );
            """
            cursor.execute(create_table_sql)
            cnxn.commit() # Valider la création de la table
            print(f"Table '{table_name}' créée avec succès ! ✅")
        else:
            print(f"La table '{table_name}' existe déjà. ⚠️")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la création de la table : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")


def Versement_Retrait(table_name,database, cnxn_str) :
    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # Requête SQL pour créer la table
        # Il est bon de vérifier si la table existe avant de la créer
        check_table_sql = f"SELECT OBJECT_ID('{table_name}', 'U')"
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                ID_VR INT IDENTITY(1,1) PRIMARY KEY,
                ID_Client INT NOT NULL,
                Date_VR DATETIME DEFAULT GETDATE(),
                Type_VR VARCHAR(255) NOT NULL,
                Debit INT NOT NULL,
                Credit INT NOT NULL,

                -- Clé secondaire
                FOREIGN KEY (ID_Client) REFERENCES Client(ID_Client)
            );
            """
            cursor.execute(create_table_sql)
            cnxn.commit() # Valider la création de la table
            print(f"Table '{table_name}' créée avec succès ! ✅")
        else:
            print(f"La table '{table_name}' existe déjà. ⚠️")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la création de la table : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")

def Tombees(table_name,database, cnxn_str) :
    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # Requête SQL pour créer la table
        # Il est bon de vérifier si la table existe avant de la créer
        check_table_sql = f"SELECT OBJECT_ID('{table_name}', 'U')"
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                ID_Tombees INT IDENTITY(1,1) PRIMARY KEY,
                ID_Client INT NOT NULL,
                DateTombees DATETIME DEFAULT GETDATE(),
                Type_Tombees VARCHAR(255) NOT NULL, -- (Intérêt, dividende, distribution FCP)
                Debit INT NOT NULL,
                Credit INT NOT NULL,

                -- Clé secondaire
                FOREIGN KEY (ID_Client) REFERENCES Client(ID_Client)
            );
            """
            cursor.execute(create_table_sql)
            cnxn.commit() # Valider la création de la table
            print(f"Table '{table_name}' créée avec succès ! ✅")
        else:
            print(f"La table '{table_name}' existe déjà. ⚠️")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la création de la table : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")

def Retenues(table_name,database, cnxn_str) :
    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # Requête SQL pour créer la table
        # Il est bon de vérifier si la table existe avant de la créer
        check_table_sql = f"SELECT OBJECT_ID('{table_name}', 'U')"
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                ID_Retenue INT IDENTITY(1,1) PRIMARY KEY,
                ID_Client INT NOT NULL,
                DateRetenue DATETIME DEFAULT GETDATE(),
                Type_Tombees VARCHAR(255) NOT NULL, -- (Commission de Valorisation & Droit de Garde)
                Debit INT NOT NULL,
                Credit INT NOT NULL,

                -- Clé secondaire
                FOREIGN KEY (ID_Client) REFERENCES Client(ID_Client)
            );
            """
            cursor.execute(create_table_sql)
            cnxn.commit() # Valider la création de la table
            print(f"Table '{table_name}' créée avec succès ! ✅")
        else:
            print(f"La table '{table_name}' existe déjà. ⚠️")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la création de la table : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")


def Portefeuille_Mouvement(table_name,database, cnxn_str) :

    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # Requête SQL pour créer la table
        # Il est bon de vérifier si la table existe avant de la créer
        check_table_sql = f"SELECT OBJECT_ID('{table_name}', 'U')"
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                ID_PortMvt INT IDENTITY(1,1) PRIMARY KEY,
                ID_Client INT NOT NULL,
                Date_Operation DATETIME DEFAULT GETDATE(),
                Type_VR VARCHAR(255) NOT NULL,
                Debit INT NOT NULL,
                Credit INT NOT NULL,

                -- Clé secondaire
                FOREIGN KEY (ID_Client) REFERENCES Client(ID_Client)
            );
            """
            cursor.execute(create_table_sql)
            cnxn.commit() # Valider la création de la table
            print(f"Table '{table_name}' créée avec succès ! ✅")
        else:
            print(f"La table '{table_name}' existe déjà. ⚠️")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la création de la table : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")


def Portefeuille_Historique(table_name,database, cnxn_str) :

    try:
        cnxn = pyodbc.connect(cnxn_str)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database}'. 🎉")

        # Requête SQL pour créer la table
        # Il est bon de vérifier si la table existe avant de la créer
        check_table_sql = f"SELECT OBJECT_ID('{table_name}', 'U')"
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if not table_exists:
            create_table_sql = f"""
            CREATE TABLE {table_name} (
                ID_PortHist INT IDENTITY(1,1) PRIMARY KEY,
                ID_Client INT NOT NULL,
                Date_Renseignement DATETIME DEFAULT GETDATE(),
                Date_Hist VARCHAR(255) NOT NULL,
                Titre_Actif VARCHAR(255) NOT NULL,
                Quantite INT NOT NULL,
                CMP FLOAT NOT NULL,
                Cours FLOAT NOT NULL,
                Interet_Couru FLOAT NOT NULL,
                Liquidite FLOAT NOT NULL,
                VALEUR_PORTEFEUILLE FLOAT NOT NULL,

            );
            """
            cursor.execute(create_table_sql)
            cnxn.commit() # Valider la création de la table
            print(f"Table '{table_name}' créée avec succès ! ✅")
        else:
            print(f"La table '{table_name}' existe déjà. ⚠️")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la création de la table : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()

    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e}")



def supprimer_base_de_donnees( db_name_to_drop, connection_string_master, database_to_connect):
    cnxn = None
    cursor = None
    try:
        cnxn = pyodbc.connect(connection_string_master)
        cursor = cnxn.cursor()
        print(f"Connexion établie à la base de données '{database_to_connect}'.")

        # Vérifier si la base de données existe
        cursor.execute(f"SELECT DB_ID('{db_name_to_drop}')")
        db_id = cursor.fetchone()[0]

        if db_id is None:
            print(f"La base de données '{db_name_to_drop}' n'existe pas. Aucune action requise.")
            return

        # Mettre la base de données en mode mono-utilisateur et tuer les connexions actives
        # C'est souvent nécessaire pour pouvoir la dropper
        print(f"Préparation à la suppression de la base de données '{db_name_to_drop}'...")
        cursor.execute(f"ALTER DATABASE [{db_name_to_drop}] SET SINGLE_USER WITH ROLLBACK IMMEDIATE;")
        print(f"Base de données '{db_name_to_drop}' mise en mode SINGLE_USER et connexions terminées.")

        # Supprimer la base de données
        drop_db_sql = f"DROP DATABASE [{db_name_to_drop}];"
        cursor.execute(drop_db_sql)
        cnxn.commit() # Valider la suppression
        print(f"Base de données '{db_name_to_drop}' supprimée avec succès ! 💥")

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erreur SQL lors de la suppression de la base de données : {sqlstate}")
        print(ex)
        if cnxn:
            cnxn.rollback()
        print("Opération de suppression annulée. ↩️")
    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {e} ❌")
    finally:
        if cursor:
            cursor.close()
        if cnxn:
            cnxn.close()
            print("Connexion fermée. 🔒")